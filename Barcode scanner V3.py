import cv2
from pyzbar.pyzbar import decode
import openpyxl
from datetime import datetime

# Function to decode barcode
def decode_barcode(frame):
    barcodes = decode(frame)
    return next(iter(barcodes), None)

# Function to export to Excel
def save_to_excel(barcode_data, pallet_num, excel_path):
    excel_file = excel_path + r"\barcode_data.xlsx"
    
    try:
        # Try to open the Excel file, if it does not exist, create a new one
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.create_sheet("Barcode Data")

    sheet = workbook["Barcode Data"]

    # Add headings if the first row has not been added yet
    if sheet.max_row == 1:
        sheet.append(["Type barcode", "Datum", "Tijd", "Pallet nummer"])

    # Add new row with barcode description, date, time and pallet number
    row = [barcode_data.data.decode('utf-8'), datetime.now().strftime("%Y-%m-%d"), datetime.now().strftime("%H:%M:%S"), pallet_num]
    sheet.append(row)

    # Save the Excel file
    workbook.save(excel_file)

# Function to start the webcam (camera 1) and display and save barcode data once
def main():
    camera_index = 0  # Use camera 1
    excel_path = r"C:\Users\janma\Desktop\Avans Hogeschool\Leerjaar 4\Minor Smart Manufacturing and Robotics\Project Palletizer\Vision Barcode"

    cap = cv2.VideoCapture(camera_index, cv2.CAP_DSHOW)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

    barcode_data = None

    while True:
        ret, frame = cap.read()

        if barcode_data is None:
            barcode_data = decode_barcode(frame)

        if barcode_data:
            # Show barcode data in live image
            cv2.putText(frame, f"Barcode: {barcode_data.data.decode('utf-8')}", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

            # Ask user to enter pallet number
            pallet_num = input("Voer het palletnummer in: ")

            # Save barcode data in Excel
            save_to_excel(barcode_data, pallet_num, excel_path)
            print("Barcode-data opgeslagen.")
            break  # Stop the loop after the barcode has been scanned and saved

        # Show live image
        cv2.imshow('Barcode Reader', frame)

        if cv2.waitKey(1) & 0xFF == 27:  # Press Esc to quit the program
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()
